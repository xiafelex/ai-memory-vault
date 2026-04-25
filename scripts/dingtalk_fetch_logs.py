#!/usr/bin/env python3
"""Fetch DingTalk work logs and export them as Markdown files."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OAPI_BASE = "https://oapi.dingtalk.com"
V1_BASE = "https://api.dingtalk.com/v1.0"
DEFAULT_START_DATE = "2024-01-01"
DEFAULT_CHUNK_DAYS = 90
DEFAULT_PAGE_SIZE = 20
DEFAULT_TIMEOUT = 30


class DingTalkAPIError(RuntimeError):
    """Raised when DingTalk returns an API-level error."""


@dataclass
class UserTarget:
    user_id: str
    department_id: str | None = None


def build_period_label(start_date: str, end_date: str) -> str:
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    if start_dt.year == end_dt.year and start_dt.month == end_dt.month:
        return start_dt.strftime("%Y-%m")
    return f"{start_dt.strftime('%Y-%m-%d')}_to_{end_dt.strftime('%Y-%m-%d')}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch DingTalk logs for one or more users and export Markdown files."
    )
    parser.add_argument("--app-key", default=os.getenv("DINGTALK_APP_KEY"))
    parser.add_argument("--app-secret", default=os.getenv("DINGTALK_APP_SECRET"))
    parser.add_argument("--excel", help="Path to an .xlsx file containing user IDs.")
    parser.add_argument("--sheet", help="Sheet name in the Excel file. Defaults to the active sheet.")
    parser.add_argument("--user-col", default="人员ID", help="Column name for user IDs in Excel.")
    parser.add_argument("--dept-col", default="部门ID", help="Column name for department IDs in Excel.")
    parser.add_argument("--user-id", help="Fetch logs for a single user.")
    parser.add_argument("--department-id", help="Department ID for a single user.")
    parser.add_argument(
        "--dept-batch-id",
        help="Fetch logs for all users in this department. Can be combined with --recursive-dept.",
    )
    parser.add_argument(
        "--recursive-dept",
        action="store_true",
        help="When used with --dept-batch-id, include users in child departments recursively.",
    )
    parser.add_argument("--start-date", default=DEFAULT_START_DATE, help="Start date, e.g. 2024-01-01.")
    parser.add_argument(
        "--end-date",
        default=datetime.now().strftime("%Y-%m-%d"),
        help="End date, e.g. 2026-04-25.",
    )
    parser.add_argument(
        "--chunk-days",
        type=int,
        default=DEFAULT_CHUNK_DAYS,
        help="How many days to fetch per request window.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
        help="Page size for /topapi/report/list.",
    )
    parser.add_argument("--template-name", help="Optional log template name filter, e.g. 日报.")
    parser.add_argument(
        "--output-dir",
        default="outputs/dingtalk_logs",
        help="Directory for generated Markdown files.",
    )
    parser.add_argument(
        "--output-org",
        help="Optional organization label. When set, files are stored under <output-dir>/<org>/<year>/<period>/.",
    )
    parser.add_argument(
        "--token-mode",
        choices=("auto", "v1", "legacy"),
        default="auto",
        help="How to obtain access token. 'auto' tries v1 first, then legacy.",
    )
    parser.add_argument(
        "--check-auth",
        action="store_true",
        help="Only validate app credentials and print a short result.",
    )
    return parser.parse_args()


def require_value(name: str, value: str | None) -> str:
    if not value:
        raise SystemExit(f"Missing required value: {name}")
    return value


def to_timestamp_ms(date_str: str, end_of_day: bool = False) -> int:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    if end_of_day:
        dt = dt + timedelta(days=1) - timedelta(milliseconds=1)
    return int(dt.timestamp() * 1000)


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", name.strip())
    return cleaned or "unknown"


def http_request(
    method: str,
    url: str,
    *,
    params: dict[str, Any] | None = None,
    data: dict[str, Any] | None = None,
    timeout: int = DEFAULT_TIMEOUT,
) -> dict[str, Any]:
    if params:
        query = urllib.parse.urlencode(params)
        url = f"{url}{'&' if '?' in url else '?'}{query}"

    body = None
    headers = {"User-Agent": "codex-dingtalk-log-fetcher/1.0"}
    if data is not None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"

    request = urllib.request.Request(url=url, method=method.upper(), data=body, headers=headers)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise DingTalkAPIError(f"HTTP {exc.code} for {url}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise DingTalkAPIError(f"Network error for {url}: {exc}") from exc

    try:
        return json.loads(payload)
    except json.JSONDecodeError as exc:
        raise DingTalkAPIError(f"Invalid JSON from {url}: {payload[:500]}") from exc


def get_access_token(app_key: str, app_secret: str, token_mode: str) -> str:
    modes = ["v1", "legacy"] if token_mode == "auto" else [token_mode]
    errors: list[str] = []

    for mode in modes:
        try:
            if mode == "v1":
                result = http_request(
                    "POST",
                    f"{V1_BASE}/oauth2/accessToken",
                    data={"appKey": app_key, "appSecret": app_secret},
                )
                token = result.get("accessToken")
            else:
                result = http_request(
                    "GET",
                    f"{OAPI_BASE}/gettoken",
                    params={"appkey": app_key, "appsecret": app_secret},
                )
                if result.get("errcode") != 0:
                    raise DingTalkAPIError(
                        f"Legacy token API failed: {result.get('errcode')} {result.get('errmsg')}"
                    )
                token = result.get("access_token")

            if token:
                return str(token)
            errors.append(f"{mode}: token missing in response")
        except DingTalkAPIError as exc:
            errors.append(f"{mode}: {exc}")

    raise DingTalkAPIError("Unable to obtain access token. " + " | ".join(errors))


def get_user_name(user_id: str, access_token: str) -> str | None:
    result = http_request(
        "GET",
        f"{OAPI_BASE}/user/get",
        params={"access_token": access_token, "userid": user_id},
    )
    if result.get("errcode") == 0:
        return result.get("name")
    return None


def get_department_name(department_id: str | None, access_token: str) -> str | None:
    if not department_id:
        return None
    result = http_request(
        "GET",
        f"{OAPI_BASE}/department/get",
        params={"access_token": access_token, "id": department_id},
    )
    if result.get("errcode") == 0:
        return result.get("name")
    return None


def list_child_departments(department_id: str, access_token: str) -> list[dict[str, Any]]:
    result = http_request(
        "POST",
        f"{OAPI_BASE}/topapi/v2/department/listsub",
        params={"access_token": access_token},
        data={"dept_id": int(department_id)},
    )
    if result.get("errcode") != 0:
        raise DingTalkAPIError(
            f"department/listsub failed for dept {department_id}: "
            f"{result.get('errcode')} {result.get('errmsg')}"
        )
    return result.get("result") or []


def list_users_in_department(department_id: str, access_token: str) -> list[UserTarget]:
    cursor = 0
    page_size = 100
    targets: list[UserTarget] = []

    while True:
        result = http_request(
            "POST",
            f"{OAPI_BASE}/topapi/v2/user/list",
            params={"access_token": access_token},
            data={
                "dept_id": int(department_id),
                "cursor": cursor,
                "size": page_size,
                "contain_access_limit": False,
            },
        )
        if result.get("errcode") != 0:
            raise DingTalkAPIError(
                f"user/list failed for dept {department_id}: "
                f"{result.get('errcode')} {result.get('errmsg')}"
            )

        payload = result.get("result") or {}
        for item in payload.get("list") or []:
            user_id = str(item.get("userid") or "").strip()
            if user_id:
                targets.append(UserTarget(user_id=user_id, department_id=department_id))

        if not payload.get("has_more"):
            return targets
        cursor = payload.get("next_cursor", 0)


def resolve_department_targets(
    department_id: str,
    access_token: str,
    recursive: bool,
) -> list[UserTarget]:
    pending = [department_id]
    seen_depts: set[str] = set()
    seen_users: set[str] = set()
    targets: list[UserTarget] = []

    while pending:
        current_dept = pending.pop(0)
        if current_dept in seen_depts:
            continue
        seen_depts.add(current_dept)

        for target in list_users_in_department(current_dept, access_token):
            if target.user_id not in seen_users:
                seen_users.add(target.user_id)
                targets.append(target)

        if recursive:
            for child in list_child_departments(current_dept, access_token):
                child_id = str(child.get("dept_id") or "").strip()
                if child_id and child_id not in seen_depts:
                    pending.append(child_id)

    return targets


def fetch_logs_for_window(
    *,
    user_id: str,
    access_token: str,
    start_ts: int,
    end_ts: int,
    page_size: int,
    template_name: str | None,
) -> list[dict[str, Any]]:
    data_list: list[dict[str, Any]] = []
    cursor: int | str = 0

    while True:
        body: dict[str, Any] = {
            "userid": user_id,
            "start_time": start_ts,
            "end_time": end_ts,
            "modified_start_time": start_ts,
            "modified_end_time": end_ts,
            "cursor": cursor,
            "size": page_size,
        }
        if template_name:
            body["template_name"] = template_name

        result = http_request(
            "POST",
            f"{OAPI_BASE}/topapi/report/list",
            params={"access_token": access_token},
            data=body,
        )

        if result.get("errcode") != 0:
            raise DingTalkAPIError(
                f"report/list failed for user {user_id}: "
                f"{result.get('errcode')} {result.get('errmsg')}"
            )

        payload = result.get("result") or {}
        data_list.extend(payload.get("data_list") or [])
        if not payload.get("has_more"):
            return data_list
        cursor = payload.get("next_cursor", 0)


def date_windows(start_date: str, end_date: str, chunk_days: int) -> Iterable[tuple[int, int, str, str]]:
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    current = start_dt

    while current <= end_dt:
        chunk_end = min(current + timedelta(days=chunk_days - 1), end_dt)
        yield (
            int(current.timestamp() * 1000),
            int((chunk_end + timedelta(days=1) - timedelta(milliseconds=1)).timestamp() * 1000),
            current.strftime("%Y-%m-%d"),
            chunk_end.strftime("%Y-%m-%d"),
        )
        current = chunk_end + timedelta(days=1)


def render_markdown(
    logs: list[dict[str, Any]],
    *,
    user_name: str,
    department_name: str,
    user_id: str,
    start_date: str,
    end_date: str,
) -> str:
    lines = [
        f"# {department_name} - {user_name} 的日志",
        "",
        f"- 人员ID: `{user_id}`",
        f"- 时间范围: `{start_date}` 至 `{end_date}`",
        f"- 日志数量: `{len(logs)}`",
        "",
    ]

    for log in sorted(logs, key=lambda item: item.get("create_time", 0)):
        create_time = log.get("create_time", 0)
        create_date = datetime.fromtimestamp(create_time / 1000).strftime("%Y-%m-%d %H:%M:%S")
        template_name = log.get("template_name") or "未命名模板"
        lines.extend(
            [
                f"## {create_date} - {template_name}",
                "",
                f"- 日志ID: `{log.get('report_id', '')}`",
                f"- 修改时间: `{datetime.fromtimestamp((log.get('modified_time', create_time)) / 1000).strftime('%Y-%m-%d %H:%M:%S')}`",
                "",
            ]
        )

        contents = log.get("contents") or []
        if not contents:
            lines.extend(["_该条日志没有正文内容。_", ""])
        else:
            for content in sorted(contents, key=lambda item: int(item.get("sort", 0))):
                key = str(content.get("key") or "未命名字段").strip()
                value = str(content.get("value") or "").strip()
                lines.extend([f"### {key}", "", value or "_空_", ""])

        lines.extend(["---", ""])

    return "\n".join(lines)


def write_markdown_file(
    content: str,
    *,
    output_dir: Path,
    department_name: str,
    user_name: str,
    user_id: str,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"{sanitize_filename(user_name)}__{sanitize_filename(user_id)}.md"
    file_path = output_dir / file_name
    file_path.write_text(content, encoding="utf-8")
    return file_path


def resolve_output_dir(args: argparse.Namespace) -> Path:
    base_dir = Path(args.output_dir).expanduser().resolve()
    if not args.output_org:
        return base_dir

    start_dt = datetime.strptime(args.start_date, "%Y-%m-%d")
    period_label = build_period_label(args.start_date, args.end_date)
    return base_dir / sanitize_filename(args.output_org) / start_dt.strftime("%Y") / period_label


def read_csv_targets(csv_path: Path, user_col: str, dept_col: str) -> list[UserTarget]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise SystemExit(f"CSV file is empty: {csv_path}")
        if user_col not in reader.fieldnames:
            raise SystemExit(f"Missing required column in CSV: {user_col}")
        if dept_col not in reader.fieldnames:
            raise SystemExit(f"Missing required column in CSV: {dept_col}")

        targets: list[UserTarget] = []
        for row in reader:
            user_value = row.get(user_col)
            dept_value = row.get(dept_col)
            if user_value in (None, ""):
                continue
            user_id = str(user_value).strip()
            department_id = str(dept_value).strip() if dept_value not in (None, "") else None
            targets.append(UserTarget(user_id=user_id, department_id=department_id))

    return targets


def read_excel_targets(excel_path: Path, sheet_name: str | None, user_col: str, dept_col: str) -> list[UserTarget]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration as exc:
        raise SystemExit(f"Excel file is empty: {excel_path}") from exc

    header_map = {name: index for index, name in enumerate(headers)}
    if user_col not in header_map:
        raise SystemExit(f"Missing required column in Excel: {user_col}")
    if dept_col not in header_map:
        raise SystemExit(f"Missing required column in Excel: {dept_col}")

    targets: list[UserTarget] = []
    for row in rows:
        user_value = row[header_map[user_col]]
        dept_value = row[header_map[dept_col]]
        if user_value in (None, ""):
            continue
        user_id = str(user_value).strip()
        department_id = str(dept_value).strip() if dept_value not in (None, "") else None
        targets.append(UserTarget(user_id=user_id, department_id=department_id))

    return targets


def resolve_targets(args: argparse.Namespace) -> list[UserTarget]:
    if args.user_id:
        return [UserTarget(user_id=str(args.user_id).strip(), department_id=args.department_id)]

    if args.dept_batch_id:
        raise SystemExit("Department targets require access token context and are resolved later.")

    if args.excel:
        excel_path = Path(args.excel).expanduser().resolve()
        if not excel_path.exists():
            raise SystemExit(f"Excel file not found: {excel_path}")
        if excel_path.suffix.lower() == ".csv":
            return read_csv_targets(excel_path, args.user_col, args.dept_col)
        return read_excel_targets(excel_path, args.sheet, args.user_col, args.dept_col)

    raise SystemExit("Provide either --user-id or --excel.")


def fetch_user_logs(target: UserTarget, args: argparse.Namespace, access_token: str, output_dir: Path) -> Path | None:
    user_name = get_user_name(target.user_id, access_token) or target.user_id
    department_name = (
        get_department_name(target.department_id, access_token) or (target.department_id or "未知部门")
    )

    print(f"[INFO] 开始处理 {department_name} / {user_name} / {target.user_id}")
    all_logs: list[dict[str, Any]] = []

    for start_ts, end_ts, start_label, end_label in date_windows(
        args.start_date, args.end_date, args.chunk_days
    ):
        print(f"[INFO] 获取区间 {start_label} -> {end_label}")
        window_logs = fetch_logs_for_window(
            user_id=target.user_id,
            access_token=access_token,
            start_ts=start_ts,
            end_ts=end_ts,
            page_size=args.page_size,
            template_name=args.template_name,
        )
        all_logs.extend(window_logs)
        time.sleep(0.2)

    if not all_logs:
        print(f"[WARN] {user_name} 在指定时间范围内未查询到日志")
        return None

    markdown = render_markdown(
        all_logs,
        user_name=user_name,
        department_name=department_name,
        user_id=target.user_id,
        start_date=args.start_date,
        end_date=args.end_date,
    )
    file_path = write_markdown_file(
        markdown,
        output_dir=output_dir,
        department_name=department_name,
        user_name=user_name,
        user_id=target.user_id,
    )
    print(f"[INFO] 已写出 {file_path}")
    return file_path


def main() -> int:
    args = parse_args()
    app_key = require_value("app_key / DINGTALK_APP_KEY", args.app_key)
    app_secret = require_value("app_secret / DINGTALK_APP_SECRET", args.app_secret)
    output_dir = resolve_output_dir(args)

    try:
        access_token = get_access_token(app_key, app_secret, args.token_mode)
    except DingTalkAPIError as exc:
        print(f"[ERROR] 获取 access token 失败: {exc}", file=sys.stderr)
        return 1

    if args.check_auth:
        print("[INFO] 凭证验证成功，access token 已获取。")
        return 0

    if args.dept_batch_id:
        targets = resolve_department_targets(
            str(args.dept_batch_id).strip(),
            access_token,
            args.recursive_dept,
        )
        dept_name = get_department_name(str(args.dept_batch_id).strip(), access_token) or args.dept_batch_id
        print(f"[INFO] 部门 {dept_name} 共解析出 {len(targets)} 名成员")
    else:
        targets = resolve_targets(args)
    if not targets:
        print("[WARN] 没有可处理的人员记录。")
        return 0

    success_count = 0
    for target in targets:
        try:
            file_path = fetch_user_logs(target, args, access_token, output_dir)
            if file_path:
                success_count += 1
        except DingTalkAPIError as exc:
            print(f"[ERROR] 处理 {target.user_id} 失败: {exc}", file=sys.stderr)

    print(f"[INFO] 完成，成功导出 {success_count} 个 Markdown 文件到 {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
